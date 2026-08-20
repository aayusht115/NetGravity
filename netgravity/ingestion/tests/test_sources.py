"""
Data source interface tests.

These pin the three assumptions the old reader made that broke on real
client data: one sheet only, fixed filenames, and folder-decides-meaning.
"""

from __future__ import annotations

import csv

import pytest

from netgravity.ingestion.sources import ErpSource, FileSource, WmsSource, discover
from netgravity.ingestion.sources.base import RecordOrigin, RecordSet


def _write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        w.writerows(rows)
    return path


def _write_xlsx(path, sheets):
    """sheets = {name: (headers, rows)}"""
    from openpyxl import Workbook
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(title=name)
        ws.append(headers)
        for r in rows:
            ws.append(r)
    wb.save(str(path))
    return path


# --- the multi-sheet gap ----------------------------------------------------

def test_every_sheet_is_read_not_just_the_first(tmp_path):
    """
    The old reader took wb.sheetnames[0] and silently dropped the rest. A
    client putting facilities on tab 1 and lanes on tab 2 lost tab 2 with no
    warning at all.
    """
    path = _write_xlsx(tmp_path / "network.xlsx", {
        "Facilities": (["Facility_ID", "Name"], [["DC_1", "Delhi"], ["DC_2", "Pune"]]),
        "Lanes":      (["Origin", "Dest"],      [["DC_1", "DC_2"]]),
        "Products":   (["SKU", "Weight"],       [["P1", 2.5]]),
    })
    sets = list(FileSource(path).record_sets())
    assert len(sets) == 3
    assert {rs.origin.sheet for rs in sets} == {"Facilities", "Lanes", "Products"}
    assert {rs.key for rs in sets} == {
        "network.xlsx#Facilities", "network.xlsx#Lanes", "network.xlsx#Products",
    }


def test_empty_sheets_are_skipped_without_failing_the_file(tmp_path):
    path = _write_xlsx(tmp_path / "wb.xlsx", {
        "Real":  (["A"], [["x"]]),
        "Blank": (["B"], []),
    })
    sets = list(FileSource(path).record_sets())
    assert [rs.origin.sheet for rs in sets] == ["Real"]


# --- filenames must not matter ---------------------------------------------

def test_any_filename_is_read(tmp_path):
    """Structured ingestion used to require the exact name 'facilities.csv'."""
    path = _write_csv(tmp_path / "Warehouse_Master_FINAL_v3.csv",
                      ["Facility_ID", "Name"], [["DC_1", "Delhi"]])
    sets = list(FileSource(path).record_sets())
    assert len(sets) == 1
    assert sets[0].row_count == 1


def test_discover_walks_nested_folders(tmp_path):
    _write_csv(tmp_path / "top.csv", ["A"], [["1"]])
    _write_csv(tmp_path / "vendor_a" / "jan.csv", ["A"], [["1"]])
    _write_csv(tmp_path / "vendor_a" / "feb.csv", ["A"], [["1"]])
    _write_csv(tmp_path / "deep" / "deeper" / "x.csv", ["A"], [["1"]])

    found = {s.path.name for s in discover(tmp_path)}
    assert found == {"top.csv", "jan.csv", "feb.csv", "x.csv"}


def test_same_sender_across_months_resolves_to_one_identity(tmp_path):
    """
    source_id is part of the memory key. If each month's file became its own
    identity, every month would look like a brand-new sender and be
    re-reviewed from scratch — so a sender's folder name wins over the
    per-file name.
    """
    _write_csv(tmp_path / "vendor_a" / "shipments_jan.csv", ["Qty"], [["10"]])
    _write_csv(tmp_path / "vendor_a" / "shipments_feb.csv", ["Qty"], [["20"]])

    ids = {s.source_id for s in discover(tmp_path)}
    assert ids == {"vendor_a"}


def test_top_level_file_falls_back_to_its_stem(tmp_path):
    _write_csv(tmp_path / "facilities.csv", ["A"], [["1"]])
    assert discover(tmp_path)[0].source_id == "facilities"


def test_non_tabular_files_are_ignored_by_discover(tmp_path):
    _write_csv(tmp_path / "real.csv", ["A"], [["1"]])
    (tmp_path / "contract.pdf").write_bytes(b"%PDF-1.4 fake")
    (tmp_path / "notes.md").write_text("hello")
    assert [s.path.name for s in discover(tmp_path)] == ["real.csv"]


# --- defensive reading ------------------------------------------------------

def test_blank_header_gets_a_positional_placeholder_not_dropped(tmp_path):
    """
    Dropping an unnamed column would shift every value to its right into the
    wrong field — silently corrupting the row.
    """
    path = _write_csv(tmp_path / "x.csv", ["ID", "", "Qty"], [["A", "ignore", "5"]])
    rs = next(iter(FileSource(path).record_sets()))
    assert rs.columns == ["ID", "column_2", "Qty"]
    assert rs.rows[0]["Qty"] == "5"


def test_unsupported_type_yields_a_warning_not_an_exception(tmp_path):
    path = tmp_path / "data.parquet"
    path.write_bytes(b"nope")
    rs = next(iter(FileSource(path).record_sets()))
    assert rs.warning is not None
    assert "unsupported" in rs.warning.lower()
    assert rs.is_empty


def test_empty_csv_is_reported_not_crashed(tmp_path):
    path = tmp_path / "empty.csv"
    path.write_text("")
    rs = next(iter(FileSource(path).record_sets()))
    assert rs.warning is not None
    assert rs.is_empty


def test_sample_rows_returns_several_rows_for_pattern_detection(tmp_path):
    """One row cannot show a pattern; pattern is what tells a shipment log
    apart from a product master when both have a 'Weight' column."""
    rows = [[f"R{i}", i] for i in range(20)]
    path = _write_csv(tmp_path / "x.csv", ["ID", "Weight"], rows)
    rs = next(iter(FileSource(path).record_sets()))
    assert len(rs.sample_rows(5)) == 5
    assert len(rs.sample_rows(3)) == 3


def test_tsv_is_read_with_tab_delimiter(tmp_path):
    path = tmp_path / "x.tsv"
    path.write_text("ID\tQty\nA\t5\n", encoding="utf-8")
    rs = next(iter(FileSource(path).record_sets()))
    assert rs.columns == ["ID", "Qty"]
    assert rs.rows[0] == {"ID": "A", "Qty": "5"}


# --- origin -----------------------------------------------------------------

def test_origin_label_reads_cleanly_with_and_without_a_sheet():
    assert RecordOrigin(container="x.csv").label == "x.csv"
    assert RecordOrigin(container="x.xlsx", sheet="Tab1").label == "x.xlsx#Tab1"


def test_record_set_reports_emptiness():
    assert RecordSet(key="k").is_empty is True
    assert RecordSet(key="k", columns=["A"], rows=[{"A": 1}]).is_empty is False


# --- the ERP stub must fail loudly -----------------------------------------

def test_erp_source_raises_instead_of_pretending():
    """
    Same discipline as the Azure OpenAI branch: an unbuilt path names the
    real gap rather than silently returning nothing, which would read as
    'the ERP had no data'.
    """
    with pytest.raises(NotImplementedError) as exc:
        list(ErpSource("sap_prod").record_sets())
    text = str(exc.value)
    assert "sap_prod" in text
    assert "not implemented" in text.lower()


def test_wms_is_a_distinct_type_sharing_the_contract():
    assert WmsSource("wms_1").source_type == "wms"
    assert ErpSource("erp_1").source_type == "erp"
    assert WmsSource("wms_1").source_id == "wms_1"
