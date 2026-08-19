"""
NetGravity — Distributor File Adapter
======================================
Handles the messy reality: every distributor sends their own spreadsheet
format — different column names, different order, different units, dates in
whatever style their ERP exports, and a free-text notes column that sometimes
hides a real surcharge.

Today somebody re-maps each of these by hand before analysis can start. This
adapter asks the model to infer the mapping, applies unit conversions, flags
anything below confidence threshold for human review rather than guessing,
and CACHES the confirmed mapping so the next file from the same distributor
needs neither a model call nor another review.

That caching is the point: the AI cost is paid once per FORMAT, not once
per file.
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from netgravity.ingestion.ai.client import get_client
from netgravity.ingestion.ai.column_mapper import propose_mapping
from netgravity.ingestion.config import IngestionConfig
from netgravity.ingestion.schemas.ingest_result import FileResult, RowIssue, Severity
from netgravity.ingestion.schemas.mapping import DistributorMapping
from netgravity.ingestion.storage.base import StorageBackend

MAPPING_CACHE_PREFIX = "distributor_mappings"
REVIEW_CONFIDENCE = 0.90


# ---------------------------------------------------------------------------
# Reading heterogeneous files
# ---------------------------------------------------------------------------

def read_rows(path: Path) -> Tuple[List[str], List[Dict[str, Any]], Optional[str]]:
    """Read CSV or Excel. Returns (columns, rows, warning)."""
    suffix = path.suffix.lower()

    if suffix in {".csv", ".txt"}:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
            return list(reader.fieldnames or []), rows, None

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return [], [], ("openpyxl is not installed — cannot read Excel files. "
                            "Run `pip install openpyxl`.")
        try:
            wb = load_workbook(str(path), data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            data = list(ws.iter_rows(values_only=True))
            if not data:
                return [], [], "spreadsheet is empty"
            headers = [str(h).strip() if h is not None else f"col_{i}"
                       for i, h in enumerate(data[0])]
            rows = [
                {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
                for r in data[1:]
                if any(v is not None for v in r)
            ]
            return headers, rows, None
        except Exception as exc:
            return [], [], f"failed to read spreadsheet: {type(exc).__name__}: {exc}"

    return [], [], f"unsupported distributor file type '{suffix}'"


# ---------------------------------------------------------------------------
# Mapping cache
# ---------------------------------------------------------------------------

def cache_key(distributor_id: str) -> str:
    return f"{MAPPING_CACHE_PREFIX}/{distributor_id}.json"


def load_cached_mapping(distributor_id: str,
                        storage: StorageBackend) -> Optional[DistributorMapping]:
    try:
        raw = storage.get_text("standardized", cache_key(distributor_id))
    except FileNotFoundError:
        return None
    try:
        return DistributorMapping.model_validate(json.loads(raw))
    except Exception:
        return None


def save_mapping(mapping: DistributorMapping, storage: StorageBackend) -> str:
    return storage.save_text(
        "standardized",
        cache_key(mapping.distributor_id),
        json.dumps(mapping.model_dump(mode="json"), indent=2, default=str),
    )


# ---------------------------------------------------------------------------
# Applying a mapping
# ---------------------------------------------------------------------------

def apply_mapping(rows: List[Dict[str, Any]], mapping: DistributorMapping,
                  file: str) -> Tuple[List[Dict[str, Any]], List[RowIssue]]:
    """
    Rewrite source rows into canonical field names, applying unit conversions.

    Conversion factors are arithmetic applied here — the model proposed the
    factor, but the multiplication is deterministic code, never a model output.
    """
    issues: List[RowIssue] = []
    rename = {m.source_column: m for m in mapping.mappings}
    out: List[Dict[str, Any]] = []

    for n, row in enumerate(rows, start=2):
        mapped: Dict[str, Any] = {}
        for source_col, value in row.items():
            spec = rename.get(source_col)
            if spec is None:
                continue    # unmapped columns are dropped, and were reported already

            if spec.conversion_factor != 1.0 and value not in (None, ""):
                try:
                    value = float(str(value).replace(",", "")) * spec.conversion_factor
                except (TypeError, ValueError):
                    issues.append(RowIssue(
                        severity=Severity.WARNING, code="R-016",
                        message=f"could not apply unit conversion "
                                f"(x{spec.conversion_factor}) to '{source_col}'",
                        source_file=file, row_number=n, column=source_col,
                        raw_value=str(value),
                    ))
            mapped[spec.target_field] = value

        if mapped:
            out.append(mapped)

    return out, issues


# ---------------------------------------------------------------------------
# Entry points
# ---------------------------------------------------------------------------

def ingest_file(path: Path, config: IngestionConfig, storage: StorageBackend,
                *, distributor_id: Optional[str] = None,
                known_ids: Optional[Set[str]] = None,
                use_cache: bool = True
                ) -> Tuple[List[Dict[str, Any]], Optional[DistributorMapping], FileResult]:
    """Ingest one distributor file, proposing or reusing a column mapping."""
    distributor_id = distributor_id or path.stem
    result = FileResult(source_file=path.name, adapter="distributor", ai_used=True)

    columns, rows, warning = read_rows(path)
    result.rows_read = len(rows)

    if warning:
        result.issues.append(RowIssue(
            severity=Severity.ERROR if not rows else Severity.WARNING,
            code="R-013", message=warning, source_file=path.name,
        ))
    if not rows:
        result.rows_rejected = result.rows_read
        result.ai_stubbed = config.stub_mode
        return [], None, result

    # --- reuse a confirmed mapping if we already have one ---
    mapping = load_cached_mapping(distributor_id, storage) if use_cache else None

    if mapping is not None and mapping.confirmed_by_human:
        result.ai_used = False
        result.ai_notes.append(
            f"reused cached mapping for '{distributor_id}' "
            f"(confirmed {mapping.created_at}) — no model call needed"
        )
    else:
        client = get_client(config)
        mapping, note = propose_mapping(
            client,
            columns=columns,
            sample_rows=rows,
            distributor_id=distributor_id,
            distributor_name=distributor_id.replace("_", " ").title(),
            known_ids=sorted(known_ids or []),
        )
        result.ai_stubbed = mapping.proposed_by == "stub"
        result.ai_notes.append(note)
        save_mapping(mapping, storage)

        # Anything the model was unsure about goes to a human — not silently trusted
        for m in mapping.needs_review:
            result.issues.append(RowIssue(
                severity=Severity.WARNING, code="R-017",
                message=f"'{m.source_column}' -> {m.target_field} at "
                        f"{m.confidence:.0%} confidence — needs human confirmation "
                        f"({m.reasoning})",
                source_file=path.name, column=m.source_column,
            ))

        for col in mapping.unmapped_columns:
            result.issues.append(RowIssue(
                severity=Severity.INFO, code="R-018",
                message=f"column '{col}' could not be mapped and was skipped",
                source_file=path.name, column=col,
            ))

    mapped_rows, issues = apply_mapping(rows, mapping, path.name)
    result.issues.extend(issues)
    result.rows_accepted = len(mapped_rows)
    result.rows_rejected = result.rows_read - result.rows_accepted

    return mapped_rows, mapping, result


def ingest_directory(distributor_dir: Path, config: IngestionConfig,
                     storage: StorageBackend,
                     known_ids: Optional[Set[str]] = None
                     ) -> Tuple[List[Dict[str, Any]], List[DistributorMapping],
                                List[FileResult]]:
    distributor_dir = Path(distributor_dir)
    all_rows: List[Dict[str, Any]] = []
    mappings: List[DistributorMapping] = []
    results: List[FileResult] = []

    for path in sorted(distributor_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx", ".xlsm"}:
            continue
        rows, mapping, result = ingest_file(path, config, storage, known_ids=known_ids)
        all_rows.extend(rows)
        if mapping is not None:
            mappings.append(mapping)
        results.append(result)

    return all_rows, mappings, results


def confirm_mapping(distributor_id: str, storage: StorageBackend) -> bool:
    """
    Mark a cached mapping as human-confirmed.

    After this, files from that distributor skip the model call entirely.
    Exposed for the future ingestion-console screen; usable from a REPL today.
    """
    mapping = load_cached_mapping(distributor_id, storage)
    if mapping is None:
        return False
    mapping.confirmed_by_human = True
    mapping.created_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    save_mapping(mapping, storage)
    return True
