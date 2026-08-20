"""
NetGravity — File Data Source
==============================
CSV and Excel, from anywhere, under any name.

WHAT CHANGED AND WHY
--------------------
The previous reader had three assumptions baked in that do not survive
contact with real client data:

  1. ONE SHEET. It read `wb.sheetnames[0]` and silently ignored the rest.
     Clients routinely send one workbook with facilities on one tab and
     lanes on another. Everything after the first tab was dropped without
     a warning.

  2. FIXED FILENAMES. Structured ingestion looked for `facilities.csv` by
     exact name. A client sending `Warehouse_Master.xlsx` got nothing, even
     though the contents were exactly what we wanted.

  3. FOLDER DECIDES MEANING. A file in `distributors/` was assumed to be
     distributor shipment data; a file in the root was assumed to be network
     master data. But a distributor can perfectly well send a facility list,
     and a client can send shipment history. Where a file sat told us
     nothing reliable about what was inside it.

This module fixes (1) and (2) by reading every sheet of every readable file
regardless of name. It does not attempt (3) at all — deciding WHAT a record
set contains is classification's job (ai/classifier.py), done from the row
data itself, not from the path.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Tuple

from netgravity.ingestion.sources.base import DataSource, RecordOrigin, RecordSet

#: Extensions we can turn into record sets. PDFs are handled by the contract
#: adapter instead — they are documents, not tables.
TABULAR_SUFFIXES = {".csv", ".tsv", ".xlsx", ".xlsm"}


def _clean_headers(raw: List[Any]) -> List[str]:
    """
    Normalise a header row.

    A blank header still occupies a column position, so it gets a positional
    placeholder rather than being dropped — dropping it would silently shift
    every value to its right into the wrong field.
    """
    headers: List[str] = []
    for i, h in enumerate(raw):
        text = str(h).strip() if h is not None else ""
        headers.append(text or f"column_{i + 1}")
    return headers


def _read_csv(path: Path, delimiter: str) -> Tuple[List[str], List[Dict[str, Any]], Optional[str]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            data = [r for r in reader]
    except OSError as exc:
        return [], [], f"could not read file: {exc}"
    except UnicodeDecodeError as exc:
        return [], [], f"file is not readable as UTF-8 text: {exc}"

    if not data:
        return [], [], "file is empty"

    headers = _clean_headers(data[0])
    rows = [
        {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        for r in data[1:]
        if any(str(v).strip() for v in r)
    ]
    return headers, rows, None


def _read_excel_sheets(path: Path) -> Tuple[List[Tuple[str, List[str], List[Dict[str, Any]]]], Optional[str]]:
    """Return [(sheet_name, headers, rows), ...] for EVERY sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], ("openpyxl is not installed — cannot read Excel files. "
                    "Run `pip install openpyxl`.")
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:
        return [], f"failed to open workbook: {type(exc).__name__}: {exc}"

    sheets: List[Tuple[str, List[str], List[Dict[str, Any]]]] = []
    for name in wb.sheetnames:
        try:
            ws = wb[name]
            data = list(ws.iter_rows(values_only=True))
        except Exception:
            continue                      # one unreadable tab must not kill the file
        if not data:
            continue
        headers = _clean_headers(list(data[0]))
        rows = [
            {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
            for r in data[1:]
            if any(v is not None and str(v).strip() for v in r)
        ]
        if rows:                          # an empty tab is not an error, just nothing
            sheets.append((name, headers, rows))
    return sheets, None


class FileSource(DataSource):
    """One file on disk. Yields one record set per sheet (CSV yields one)."""

    source_type = "file"

    def __init__(self, path: Path, source_id: Optional[str] = None):
        self.path = Path(path)
        self._source_id = source_id or self.path.stem

    @property
    def source_id(self) -> str:
        return self._source_id

    def record_sets(self) -> Iterator[RecordSet]:
        suffix = self.path.suffix.lower()

        if suffix in {".csv", ".tsv"}:
            headers, rows, warning = _read_csv(
                self.path, delimiter="\t" if suffix == ".tsv" else ",")
            yield RecordSet(
                key=self.path.name,
                columns=headers,
                rows=rows,
                origin=RecordOrigin(source_type=self.source_type,
                                    source_id=self._source_id,
                                    container=self.path.name),
                warning=warning,
            )
            return

        if suffix in {".xlsx", ".xlsm"}:
            sheets, warning = _read_excel_sheets(self.path)
            if warning:
                yield RecordSet(
                    key=self.path.name,
                    origin=RecordOrigin(source_type=self.source_type,
                                        source_id=self._source_id,
                                        container=self.path.name),
                    warning=warning,
                )
                return
            for sheet_name, headers, rows in sheets:
                yield RecordSet(
                    key=f"{self.path.name}#{sheet_name}",
                    columns=headers,
                    rows=rows,
                    origin=RecordOrigin(source_type=self.source_type,
                                        source_id=self._source_id,
                                        container=self.path.name,
                                        sheet=sheet_name),
                )
            return

        yield RecordSet(
            key=self.path.name,
            origin=RecordOrigin(source_type=self.source_type,
                                source_id=self._source_id,
                                container=self.path.name),
            warning=f"unsupported file type '{suffix}'",
        )


def discover(root: Path) -> List[FileSource]:
    """
    Find every tabular file under `root`, at any depth.

    SOURCE IDENTITY. A file sitting in a subfolder takes that FOLDER's name
    as its source id; a file at the top level takes its own stem. That is
    not cosmetic — source_id is one of the things confirmed mappings are
    keyed on, so it has to be stable for the same sender across files. A
    vendor sending shipments_jan.xlsx and shipments_feb.xlsx must resolve to
    ONE identity, or every month looks like a brand-new sender and gets
    re-reviewed from scratch. Folder-per-sender gives us that; bare filenames
    do not.
    """
    root = Path(root)
    if root.is_file():
        return [FileSource(root)]

    sources: List[FileSource] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in TABULAR_SUFFIXES:
            continue
        relative = path.relative_to(root)
        source_id = relative.parts[0] if len(relative.parts) > 1 else path.stem
        sources.append(FileSource(path, source_id=source_id))
    return sources
