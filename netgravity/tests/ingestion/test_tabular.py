"""
Unified tabular path tests.

The headline claim: this replaces two readers with different capabilities
without changing what reaches the optimiser. That is proved by comparing the
content-addressed data_version of both paths — identical inputs must produce
an identical hash.
"""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

from netgravity.ingestion import tabular
from netgravity.ingestion.adapters import structured
from netgravity.ingestion.builder import build_network, summarise
from netgravity.ingestion.config import IngestionConfig
from netgravity.ingestion.memory import FieldMemory
from netgravity.ingestion.pipeline import run_ingestion
from netgravity.ingestion.schemas.content import ContentType
from netgravity.ingestion.storage.local import LocalStorage

MOCK_DIR = Path(__file__).resolve().parents[3] / "data" / "mock" / "india"


@pytest.fixture
def offline_config():
    """No key: the rules-only path, so nothing can reach a network."""
    config = IngestionConfig()
    config.llm_api_key = None
    return config


@pytest.fixture
def storage(tmp_path):
    return LocalStorage(tmp_path)


def _write_csv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        writer.writerows(rows)
    return path


# --- equivalence with the path it replaces ---------------------------------

def test_the_unified_path_produces_an_identical_network(offline_config, storage):
    """
    data_version is a content hash of the assembled network. Identical
    hashes mean the two paths did not merely produce similar results — they
    produced the same bytes.
    """
    legacy = structured.ingest_directory(MOCK_DIR)
    legacy_network, _ = build_network(
        facilities=legacy.facilities, products=legacy.products,
        demands=legacy.demands, lanes=legacy.lanes, network_id="n")

    outcome = tabular.ingest_tabular(MOCK_DIR, offline_config, storage,
                                     auto_confirm=True)
    parsed = tabular.parse_into_records(outcome)
    unified_network, _ = build_network(
        facilities=parsed["facilities"], products=parsed["products"],
        demands=parsed["demands"], lanes=parsed["lanes"], network_id="n")

    assert unified_network.data_version == legacy_network.data_version
    assert summarise(unified_network) == summarise(legacy_network)


def test_the_pipeline_agrees_with_itself_across_both_paths(offline_config):
    legacy = run_ingestion(MOCK_DIR, config=offline_config, save=False)
    unified = run_ingestion(MOCK_DIR, config=offline_config, save=False,
                            unified=True, auto_confirm=True)

    assert unified.ok is True
    assert unified.report.data_version == legacy.report.data_version
    assert unified.report.counts == legacy.report.counts
    assert unified.report.engine_validation_passed


# --- what the old readers could not do -------------------------------------

def test_every_sheet_of_a_workbook_is_ingested(tmp_path, offline_config, storage):
    """The old reader took the first sheet and dropped the rest silently."""
    from openpyxl import Workbook

    path = tmp_path / "client_master.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    facilities = workbook.create_sheet("Sites")
    facilities.append(["Facility_ID", "Facility_Name", "Type", "Capacity_Units"])
    facilities.append(["DC_1", "Delhi DC", "DC", 5000])
    products = workbook.create_sheet("SKUs")
    products.append(["Product_ID", "Product_Name", "Weight", "Unit_Value"])
    products.append(["P1", "Widget", 2.5, 100])
    workbook.save(str(path))

    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage,
                                     auto_confirm=True)
    types = {m.content_type for m in outcome.mappings}
    assert ContentType.FACILITY in types
    assert ContentType.PRODUCT in types


def test_a_file_with_an_unexpected_name_is_still_read(tmp_path, offline_config,
                                                      storage):
    """Structured ingestion used to require the literal name facilities.csv."""
    _write_csv(tmp_path / "Warehouse_Master_FINAL.csv",
               ["Facility_ID", "Facility_Name", "Type", "Capacity_Units"],
               [["DC_1", "Delhi", "DC", 5000]])
    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage,
                                     auto_confirm=True)
    assert outcome.mappings[0].content_type == ContentType.FACILITY
    assert outcome.network_rows.get(ContentType.FACILITY)


# --- routing is by content, not by folder ----------------------------------

def test_shipment_data_goes_to_staging_even_from_the_root(tmp_path,
                                                          offline_config, storage):
    _write_csv(tmp_path / "anything.csv",
               ["Location Code", "Qty", "Despatch Dt", "Vehicle No"],
               [["MKT_A", "10", "2026-01-05", "MH01AB1234"]])
    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage,
                                     auto_confirm=True)
    mapping = outcome.mappings[0]
    assert mapping.content_type == ContentType.SHIPMENT_LOG
    assert mapping.destination == "staging"
    assert not outcome.network_rows


def test_facility_data_inside_a_distributors_folder_still_feeds_the_network(
        tmp_path, offline_config, storage):
    """A distributor can send a facility list. The folder is not evidence."""
    _write_csv(tmp_path / "distributors" / "sites.csv",
               ["Facility_ID", "Facility_Name", "Type", "Capacity_Units"],
               [["DC_9", "Pune", "DC", 1000]])
    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage,
                                     auto_confirm=True)
    assert outcome.mappings[0].destination == "network"
    assert outcome.network_rows.get(ContentType.FACILITY)


def test_an_unidentifiable_file_is_held_not_routed_on_a_guess(tmp_path,
                                                              offline_config,
                                                              storage):
    _write_csv(tmp_path / "mystery.csv", ["zzz", "qqq"], [["1", "2"]])
    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage)
    assert outcome.held
    assert outcome.held[0].content_type == ContentType.UNKNOWN
    assert not outcome.network_rows and not outcome.staging_rows


# --- nothing unconfirmed reaches the optimiser -----------------------------

def test_pending_optimiser_bound_columns_are_not_applied(tmp_path,
                                                         offline_config, storage):
    """
    Without auto_confirm, network-bound mappings are held. This is the
    "confirm before it touches the optimiser" rule, and it must actually
    hold — not merely be flagged while the rows go through anyway.
    """
    _write_csv(tmp_path / "sites.csv",
               ["Facility_ID", "Facility_Name", "Type"], [["DC_1", "Delhi", "DC"]])
    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage)

    mapping = outcome.mappings[0]
    assert mapping.pending, "network columns should await confirmation"
    assert mapping.rename_map == {}
    assert not outcome.network_rows.get(ContentType.FACILITY)


def test_auto_confirm_settles_them_and_says_it_was_a_machine(tmp_path,
                                                             offline_config,
                                                             storage):
    """The audit trail must never claim a person looked when nobody did."""
    _write_csv(tmp_path / "sites.csv",
               ["Facility_ID", "Facility_Name", "Type"], [["DC_1", "Delhi", "DC"]])
    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage,
                                     auto_confirm=True)

    assert outcome.mappings[0].pending == []
    assert outcome.network_rows.get(ContentType.FACILITY)

    observations = FieldMemory(storage).observations_for("FACILITY", "Facility_ID")
    assert observations
    assert observations[0].confirmed_by == "auto"


def test_confirmations_are_remembered_for_the_next_run(tmp_path, offline_config,
                                                       storage):
    _write_csv(tmp_path / "sites.csv",
               ["Facility_ID", "Facility_Name", "Type"], [["DC_1", "Delhi", "DC"]])
    tabular.ingest_tabular(tmp_path, offline_config, storage, auto_confirm=True)

    second = tabular.ingest_tabular(tmp_path, offline_config, storage)
    assert second.mappings[0].pending == []
    assert second.network_rows.get(ContentType.FACILITY)


# --- reporting --------------------------------------------------------------

def test_the_run_reports_what_is_awaiting_review(offline_config):
    result = run_ingestion(MOCK_DIR, config=offline_config, save=False,
                           unified=True)
    request = result.review_request
    assert not request.is_empty
    assert request.run_id == result.report.run_id
    assert all(item.question for item in request.items)


def test_a_legacy_run_exposes_an_empty_review_request(offline_config):
    """The property must be safe to call whichever path ran."""
    result = run_ingestion(MOCK_DIR, config=offline_config, save=False)
    assert result.review_request.is_empty


def test_staging_rows_are_written_to_the_staging_zone(offline_config, storage):
    outcome = tabular.ingest_tabular(MOCK_DIR, offline_config, storage,
                                     auto_confirm=True)
    written = tabular.save_staging(outcome, storage, "india")
    assert written
    assert any("shipment_log" in w for w in written)


def test_an_unreadable_file_is_reported_not_fatal(tmp_path, offline_config,
                                                  storage):
    (tmp_path / "broken.csv").write_bytes(b"\xff\xfe\x00binary")
    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage)
    assert outcome.results
    assert any(i.code == "R-024" for r in outcome.results for i in r.issues)


def test_an_empty_directory_is_not_an_error(tmp_path, offline_config, storage):
    outcome = tabular.ingest_tabular(tmp_path, offline_config, storage)
    assert outcome.mappings == []
    assert outcome.needs_review is False


def test_the_legacy_distributor_adapter_does_not_double_ingest(offline_config):
    """
    The unified path already read the distributors folder. Running the old
    adapter as well would ingest the same rows twice.
    """
    result = run_ingestion(MOCK_DIR, config=offline_config, save=False,
                           unified=True, auto_confirm=True)
    assert result.distributor_mappings == []
